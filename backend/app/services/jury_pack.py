"""Сборка ZIP с Excel по каждому кейсу (команды, эксперты кейса, финальные баллы)."""

from __future__ import annotations

import io
import re
import zipfile
from datetime import datetime, timezone
from typing import Iterable

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.criterion import Criterion
from app.models.project_case import ProjectCase, ProjectCaseExpert, ProjectCaseTeam
from app.models.score import Score
from app.models.team import Team
from app.models.user import User


def _case_lookup_merged(cases: Iterable[ProjectCase]) -> dict[int, ProjectCase]:
    """Ordinal и id кейса — в teams.case_number могли хранить любое из двух."""
    m: dict[int, ProjectCase] = {}
    for c in cases:
        m[int(c.ordinal)] = c
        m[int(c.id)] = c
    return m


def _team_solution_link(team: Team) -> str:
    """В интерфейсе «решение» = solution_submission_url; repo_url — запасной/регистрация."""
    sub = (team.solution_submission_url or "").strip()
    if sub:
        return sub
    return (team.repo_url or "").strip()


def _safe_filename_part(title: str, max_len: int = 48) -> str:
    t = re.sub(r"[^\w\s\-]", "", title, flags=re.UNICODE)
    t = re.sub(r"\s+", "_", t.strip())
    t = t.strip("-_") or "case"
    return t[:max_len]


def case_for_team(team: Team, case_by_key: dict[int, ProjectCase]) -> ProjectCase | None:
    if team.case_assignment and team.case_assignment.case:
        return team.case_assignment.case
    if team.case_number is None:
        return None
    return case_by_key.get(int(team.case_number))


def teams_for_case(case: ProjectCase, all_teams: Iterable[Team], case_by_key: dict[int, ProjectCase]) -> list[Team]:
    out: list[Team] = []
    for t in all_teams:
        c = case_for_team(t, case_by_key)
        if c is not None and c.id == case.id:
            out.append(t)
    return sorted(out, key=lambda x: x.id)


def unassigned_teams(all_teams: Iterable[Team], case_by_key: dict[int, ProjectCase]) -> list[Team]:
    return sorted([t for t in all_teams if case_for_team(t, case_by_key) is None], key=lambda x: x.id)


def _write_workbook(
    teams: list[Team],
    case_experts: list[ProjectCaseExpert],
    score_rows: list[tuple],
    case_by_key: dict[int, ProjectCase],
) -> bytes:
    wb = Workbook()
    ws_t = wb.active
    ws_t.title = "Teams"
    ws_t.append(["team_id", "name", "case_number", "repo_url"])
    for t in teams:
        rc = case_for_team(t, case_by_key)
        case_ord = rc.ordinal if rc else None
        ws_t.append([t.id, t.name, case_ord, _team_solution_link(t)])

    ws_e = wb.create_sheet("CaseExperts")
    ws_e.append(["user_id", "username", "email"])
    for link in case_experts:
        u = link.user
        ws_e.append([str(u.id), u.username, u.email])

    ws_s = wb.create_sheet("Scores_final")
    ws_s.append(
        [
            "team_id",
            "team_name",
            "expert_username",
            "criterion_name",
            "max_score",
            "value",
            "is_final",
            "jury_comment",
        ]
    )
    for row in score_rows:
        ws_s.append(list(row))

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


async def _fetch_score_rows(db: AsyncSession, team_ids: list[int]) -> list[tuple]:
    if not team_ids:
        return []
    stmt = (
        select(
            Team.id,
            Team.name,
            User.username,
            Criterion.name,
            Criterion.max_score,
            Score.value,
            Score.is_final,
            Score.jury_comment,
        )
        .select_from(Score)
        .join(Team, Team.id == Score.team_id)
        .join(User, User.id == Score.expert_id)
        .join(Criterion, Criterion.id == Score.criterion_id)
        .where(Score.team_id.in_(team_ids), Score.is_final.is_(True))
        .order_by(Team.id, User.username, Criterion.id)
    )
    r = await db.execute(stmt)
    return [
        (
            int(a),
            str(b),
            str(c),
            str(d),
            float(e),
            float(f),
            bool(g),
            (str(h) if h else ""),
        )
        for a, b, c, d, e, f, g, h in r.all()
    ]


async def build_jury_zip_bytes(db: AsyncSession) -> tuple[bytes, str]:
    cases = (
        (
            await db.execute(
                select(ProjectCase)
                .order_by(ProjectCase.ordinal.asc())
                .options(
                    selectinload(ProjectCase.teams).selectinload(ProjectCaseTeam.team),
                    selectinload(ProjectCase.experts).selectinload(ProjectCaseExpert.user),
                )
            )
        )
        .scalars()
        .all()
    )
    case_by_key = _case_lookup_merged(cases)

    all_teams = (
        (await db.execute(select(Team).options(selectinload(Team.case_assignment).selectinload(ProjectCaseTeam.case))))
        .scalars()
        .all()
    )

    zip_buf = io.BytesIO()
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for case in cases:
            teams = teams_for_case(case, all_teams, case_by_key)
            team_ids = [t.id for t in teams]
            score_rows = await _fetch_score_rows(db, team_ids)
            xlsx = _write_workbook(teams, list(case.experts), score_rows, case_by_key)
            fname = f"case_{case.ordinal}_{_safe_filename_part(case.title)}.xlsx"
            zf.writestr(fname, xlsx)

        loose = unassigned_teams(all_teams, case_by_key)
        if loose:
            team_ids = [t.id for t in loose]
            score_rows = await _fetch_score_rows(db, team_ids)
            xlsx = _write_workbook(loose, [], score_rows, case_by_key)
            zf.writestr("case_unassigned.xlsx", xlsx)

    zip_buf.seek(0)
    return zip_buf.getvalue(), f"jury_pack_{stamp}.zip"
